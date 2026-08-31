#!/usr/bin/env python3
"""Render a single row (by rank) from a content-system xlsx into a video."""

import argparse
import sys

import openpyxl

from pipeline.assemble import assemble_video
from pipeline.config import Config
from pipeline.providers.images import get_image_provider
from pipeline.providers.tts import get_tts_provider
from pipeline.script_builder import build_script, read_row


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sheet", required=True)
    p.add_argument("--rank", required=True, type=int)
    p.add_argument("--video-dir", required=True)
    p.add_argument("--build-dir", default=None, help="Defaults to <video-dir>/_build/<rank>_<slug>")
    p.add_argument("--image-dir", default=None,
                   help="Folder of your own images (required by --image-provider local)")
    p.add_argument("--fit", choices=["cover", "contain"], default=None,
                   help="cover = fill and center-crop (default); contain = fit whole image "
                        "with a blurred backdrop, nothing cropped")
    p.add_argument("--allow-missing-images", action="store_true",
                   help="With --image-provider local, fill any missing shot with placeholder "
                        "art instead of erroring")
    p.add_argument("--image-provider", default=None, help="placeholder | openai (defaults to $IMAGE_PROVIDER)")
    p.add_argument("--tts-provider", default=None, help="espeak | openai | elevenlabs (defaults to $TTS_PROVIDER)")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.sheet)
    ws = wb["Content System"] if "Content System" in wb.sheetnames else wb.active

    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == args.rank:
            target_row = r
            break
    if target_row is None:
        print(f"No row with Rank == {args.rank} found in {args.sheet}")
        sys.exit(1)

    row = read_row(ws, target_row)
    script = build_script(row)

    config = Config()
    config.check_dependencies()
    if args.fit:
        config.FIT_MODE = args.fit
    image_provider = get_image_provider(
        args.image_provider or config.IMAGE_PROVIDER, config,
        image_dir=args.image_dir, allow_missing=args.allow_missing_images)
    tts_provider = get_tts_provider(args.tts_provider or config.TTS_PROVIDER, config)

    build_dir = args.build_dir or f"{args.video_dir}/_build/{script.rank:02d}_{script.slug}"

    print(f"Rendering rank #{script.rank}: {script.title}")
    print(f"  narration ({len(script.full_text.split())} words): {script.full_text}")
    print(f"  {len(script.image_prompts)} image(s) planned")

    result = assemble_video(script, image_provider, tts_provider, build_dir, args.video_dir, config)

    print(f"  duration: {result['duration']:.1f}s")
    print(f"  video:  {result['video_path']}")
    print(f"  thumb:  {result['thumb_path']}")


if __name__ == "__main__":
    main()
