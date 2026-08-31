#!/usr/bin/env python3
"""
Renders videos for every not-yet-rendered row in a content-system xlsx, in
rank order, writing output files named to match what post_daily.py expects
(<rank>_<slug>.mp4 in --video-dir) and marking each row as Rendered so a
second run picks up where the first left off.

Run this ahead of post_daily.py -- e.g. render a week's worth in one batch
with --limit 7, then let post_daily.py post one a day from what's ready.
"""

import argparse
import datetime
import sys

import openpyxl

from pipeline.assemble import assemble_video
from pipeline.config import Config
from pipeline.providers.images import get_image_provider
from pipeline.providers.tts import get_tts_provider
from pipeline.script_builder import build_script, read_row

RENDERED_COL = 24
RENDERED_DATE_COL = 25


def ensure_tracking_columns(ws):
    if ws.cell(1, RENDERED_COL).value in (None, ""):
        ws.cell(1, RENDERED_COL, "Rendered")
    if ws.cell(1, RENDERED_DATE_COL).value in (None, ""):
        ws.cell(1, RENDERED_DATE_COL, "Rendered Date (UTC)")


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--sheet", required=True)
    p.add_argument("--video-dir", required=True)
    p.add_argument("--limit", type=int, default=1, help="Max number of videos to render this run")
    p.add_argument("--image-dir", default=None,
                   help="Folder of your own images (required by --image-provider local)")
    p.add_argument("--fit", choices=["cover", "contain"], default=None,
                   help="cover = fill and center-crop (default); contain = fit whole image "
                        "with a blurred backdrop, nothing cropped")
    p.add_argument("--allow-missing-images", action="store_true",
                   help="With --image-provider local, fill any missing shot with placeholder "
                        "art instead of erroring")
    p.add_argument("--image-provider", default=None)
    p.add_argument("--tts-provider", default=None)
    p.add_argument("--stop-on-error", action="store_true", help="Default is to log and continue to the next idea")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.sheet)
    ws = wb["Content System"] if "Content System" in wb.sheetnames else wb.active
    ensure_tracking_columns(ws)

    config = Config()
    config.check_dependencies()
    if args.fit:
        config.FIT_MODE = args.fit
    image_provider = get_image_provider(
        args.image_provider or config.IMAGE_PROVIDER, config,
        image_dir=args.image_dir, allow_missing=args.allow_missing_images)
    tts_provider = get_tts_provider(args.tts_provider or config.TTS_PROVIDER, config)

    candidates = []
    for r in range(2, ws.max_row + 1):
        rank = ws.cell(r, 1).value
        title = ws.cell(r, 2).value
        if not rank or not title:
            continue
        if ws.cell(r, RENDERED_COL).value == "Rendered":
            continue
        candidates.append(r)
    candidates.sort(key=lambda r: ws.cell(r, 1).value)

    if not candidates:
        print("Nothing left to render -- every row is already marked Rendered.")
        sys.exit(0)

    rendered_count = 0
    for r in candidates:
        if rendered_count >= args.limit:
            break
        row = read_row(ws, r)
        script = build_script(row)
        build_dir = f"{args.video_dir}/_build/{script.rank:02d}_{script.slug}"

        print(f"[{rendered_count + 1}/{args.limit}] Rendering rank #{script.rank}: {script.title}")
        try:
            result = assemble_video(script, image_provider, tts_provider, build_dir, args.video_dir, config)
        except Exception as exc:
            print(f"  FAILED: {exc}")
            if args.stop_on_error:
                raise
            continue

        ws.cell(r, RENDERED_COL, "Rendered")
        ws.cell(r, RENDERED_DATE_COL, datetime.datetime.utcnow().isoformat(timespec="seconds"))
        wb.save(args.sheet)  # save after each video so partial batches aren't lost

        print(f"  duration: {result['duration']:.1f}s -> {result['video_path']}")
        rendered_count += 1

    print(f"Rendered {rendered_count} video(s) this run.")


if __name__ == "__main__":
    main()
