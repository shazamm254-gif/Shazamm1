#!/usr/bin/env python3
"""
Builds the folder skeleton for supplying your own images, plus a manifest
describing what belongs in every slot.

Run this before rendering with --image-provider local. For each idea it
creates one folder named exactly what the renderer will look for, and
writes a CSV listing every shot: its narration line, the camera move that
will be applied to it, and the exact filename to drop in.

  python3 list_shots.py --sheet MySheet.xlsx --image-dir ./my_images
  python3 list_shots.py --sheet MySheet.xlsx --image-dir ./my_images --rank 1

Then put your images in, and render:

  python3 render_one.py --sheet MySheet.xlsx --rank 1 \\
      --video-dir ./videos --image-provider local --image-dir ./my_images

Anything still missing is reported by --check rather than discovered
halfway through a batch render.
"""

import argparse
import csv
import glob
import os
import sys

import openpyxl

from pipeline.script_builder import build_script, read_row

EXTS = ("jpg", "jpeg", "png", "webp", "bmp", "tif", "tiff", "JPG", "JPEG", "PNG", "WEBP")

MOTION_NOTES = {
    "push_in": "slow zoom IN - leave a little headroom, the edges get cropped",
    "pull_out": "slow zoom OUT - starts tight, so keep the subject centered",
    "pan_left_right": "pans LEFT to RIGHT - a wider image gives the pan room to travel",
    "pan_right_left": "pans RIGHT to LEFT - a wider image gives the pan room to travel",
    "static_drift": "almost still - fine for a detailed image you want read properly",
    "hard_zoom_then_push_in": "fast punch-in then slow push - put the subject dead center",
}


def find_existing(folder, stem):
    for ext in EXTS:
        hits = sorted(glob.glob(os.path.join(folder, f"{stem}.{ext}")))
        if hits:
            return hits[0]
    return None


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--sheet", required=True)
    p.add_argument("--image-dir", required=True)
    p.add_argument("--rank", type=int, default=None,
                   help="Only this rank; default is every row in the sheet")
    p.add_argument("--check", action="store_true",
                   help="Only report which slots are still empty; create nothing")
    p.add_argument("--manifest", default=None,
                   help="CSV path (default: <image-dir>/shot_manifest.csv)")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.sheet)
    ws = wb["Content System"] if "Content System" in wb.sheetnames else wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        rank = ws.cell(r, 1).value
        if not rank or not ws.cell(r, 2).value:
            continue
        if args.rank is not None and rank != args.rank:
            continue
        rows.append(r)
    if not rows:
        print(f"No matching rows found in {args.sheet}")
        sys.exit(1)

    manifest_path = args.manifest or os.path.join(args.image_dir, "shot_manifest.csv")
    records = []
    total_slots = missing_slots = 0

    for r in rows:
        script = build_script(read_row(ws, r))
        folder_name = f"{script.rank:02d}_{script.slug}"
        folder = os.path.join(args.image_dir, folder_name)
        if not args.check:
            os.makedirs(folder, exist_ok=True)

        for i, (beat, shot) in enumerate(zip(script.beats, script.image_prompts)):
            total_slots += 1
            existing = find_existing(folder, f"shot_{i:02d}") if os.path.isdir(folder) else None
            if existing is None:
                missing_slots += 1
            records.append({
                "rank": script.rank,
                "title": script.title,
                "folder": folder_name,
                "filename": f"shot_{i:02d}.jpg",
                "narration_line": beat.caption,
                "camera_move": shot.motion,
                "framing_note": MOTION_NOTES.get(shot.motion, ""),
                "shot_description": shot.prompt,
                "have_it": "yes" if existing else "NO",
            })

        # Thumbnail slot (optional -- falls back to shot_00).
        total_slots += 1
        thumb_existing = (find_existing(folder, "thumb") if os.path.isdir(folder) else None)
        if thumb_existing is None:
            missing_slots += 1
        records.append({
            "rank": script.rank,
            "title": script.title,
            "folder": folder_name,
            "filename": "thumb.jpg",
            "narration_line": "(thumbnail - optional, falls back to shot_00)",
            "camera_move": "none",
            "framing_note": "still image, no motion applied",
            "shot_description": script.thumbnail_prompt,
            "have_it": "yes" if thumb_existing else "optional",
        })

    if not args.check:
        os.makedirs(args.image_dir, exist_ok=True)
        with open(manifest_path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=list(records[0].keys()))
            w.writeheader()
            w.writerows(records)

    required_missing = sum(1 for r in records if r["have_it"] == "NO")
    print(f"ideas: {len(rows)}   shot slots: {total_slots - len(rows)} required "
          f"+ {len(rows)} optional thumbnails")
    print(f"required images still missing: {required_missing}")
    if not args.check:
        print(f"folders created under: {args.image_dir}")
        print(f"manifest: {manifest_path}")
    if required_missing and args.check:
        by_folder = {}
        for rec in records:
            if rec["have_it"] == "NO":
                by_folder.setdefault(rec["folder"], []).append(rec["filename"])
        for folder, files in sorted(by_folder.items())[:20]:
            print(f"  {folder}: missing {len(files)} -> {', '.join(files[:4])}"
                  f"{' ...' if len(files) > 4 else ''}")
        if len(by_folder) > 20:
            print(f"  ... and {len(by_folder) - 20} more folders")


if __name__ == "__main__":
    main()
