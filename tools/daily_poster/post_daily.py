#!/usr/bin/env python3
"""
Posts one video to YouTube per run, pulling metadata from a content-system
xlsx (the same schema used across the Scam Mechanics, Financial Dark
Patterns, Gym Rats, and Designed to Trick You sheets) and tracking what's
been posted back into that same file.

Intended usage: point a daily cron job at this script. Each run finds the
highest-ranked idea that (a) hasn't been posted yet and (b) has a matching
video file sitting in --video-dir, uploads it, and marks the row as posted.
If no ready video is found, it exits cleanly without posting -- it never
invents a video or posts out of order.

Setup (one time):
  1. Google Cloud Console -> new project -> enable "YouTube Data API v3".
  2. OAuth consent screen -> External -> add yourself as a test user.
  3. Credentials -> Create OAuth client ID -> Application type: Desktop app.
     Download the JSON, save it as client_secret.json.
  4. pip install -r requirements.txt
  5. First run (use --privacy private and --dry-run first):
       python3 post_daily.py --sheet system.xlsx --video-dir ./videos \
           --client-secrets client_secret.json --token token.json \
           --privacy private --dry-run
     This opens a browser once to authorize; token.json is cached after
     that and reused silently on every future run.
  6. Once you're happy with a real (non-dry-run) private test upload,
     switch --privacy to public and wire the script into cron.

Video file convention: name rendered files "<rank>_<slug>.mp4", which is
exactly what the render pipeline writes -- e.g.
"01_why_the_milk_is_always_at_the_back_of_the_store.mp4". An optional
per-row filename in the "VideoFile" column (column W) overrides the
convention if you'd rather name files freely.

Thumbnail convention (optional): "<rank>_<slug>_thumb.jpg" (also what the
render pipeline writes) or a bare "<rank>_thumb.jpg" in the same folder
gets set as the custom thumbnail if present.

Quota note: one upload costs ~1600 units against the default 10,000/day
quota, so one post a day leaves headroom to run this for multiple channels
from the same Google Cloud project if needed (each channel still needs its
own --token file, since each token is tied to one authorized account).
"""

import argparse
import datetime
import glob
import os
import re
import sys
import unicodedata

import openpyxl

SCOPES = ["https://www.googleapis.com/auth/youtube.upload"]

STATUS_COL = 20        # T
POSTED_DATE_COL = 21   # U
VIDEO_ID_COL = 22      # V
VIDEO_FILE_COL = 23    # W

TITLE_COL = 2
HOOK_COL = 4
WHY_END_COL = 6
CATEGORY_COL = 3
SEARCH_INTENT_COL = 16
TWIST_COL = 18


def slugify(text):
    # Mirrors render_pipeline/pipeline/script_builder.py's slugify() exactly
    # so filenames line up between the two tools without a hard dependency
    # between them.
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    return text[:60]


def load_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        rank = ws.cell(r, 1).value
        title = ws.cell(r, TITLE_COL).value
        if not rank or not title:
            continue
        rows.append(
            {
                "row": r,
                "rank": int(rank),
                "title": title,
                "category": ws.cell(r, CATEGORY_COL).value or "",
                "hook": ws.cell(r, HOOK_COL).value or "",
                "why_end": ws.cell(r, WHY_END_COL).value or "",
                "search_intent": ws.cell(r, SEARCH_INTENT_COL).value or "",
                "twist": ws.cell(r, TWIST_COL).value or "",
                "status": ws.cell(r, STATUS_COL).value,
                "video_file_override": ws.cell(r, VIDEO_FILE_COL).value,
            }
        )
    rows.sort(key=lambda x: x["rank"])
    return rows


def ensure_tracking_columns(ws):
    headers = {
        STATUS_COL: "Status",
        POSTED_DATE_COL: "Posted Date (UTC)",
        VIDEO_ID_COL: "YouTube Video ID",
        VIDEO_FILE_COL: "VideoFile (optional override)",
    }
    for col, label in headers.items():
        if ws.cell(1, col).value in (None, ""):
            ws.cell(1, col, label)


def find_video_file(video_dir, rank, title, override):
    if override:
        candidate = os.path.join(video_dir, override)
        return candidate if os.path.isfile(candidate) else None
    # Matching on rank alone isn't enough once more than one content-system
    # sheet shares a --video-dir -- every sheet has its own "rank 1", so a
    # bare "01_*.mp4" glob could just as easily return a different idea's
    # video entirely. Requiring the title's slug as a prefix scopes the
    # match to this specific idea.
    slug = slugify(title)
    for ext in ("mp4", "mov", "MP4", "MOV"):
        matches = glob.glob(os.path.join(video_dir, f"{rank:02d}_{slug}*.{ext}"))
        if matches:
            return matches[0]
    return None


def find_thumbnail(video_dir, rank, title):
    slug = slugify(title)
    for ext in ("jpg", "jpeg", "png", "JPG", "PNG"):
        matches = glob.glob(os.path.join(video_dir, f"{rank:02d}_{slug}*thumb.{ext}"))
        if matches:
            return matches[0]
    # Fall back to the older bare "<rank>_thumb.jpg" convention.
    for ext in ("jpg", "jpeg", "png", "JPG", "PNG"):
        matches = glob.glob(os.path.join(video_dir, f"{rank:02d}_thumb.{ext}"))
        if matches:
            return matches[0]
    return None


def build_metadata(idea, category_id):
    title = idea["title"].strip()

    description_parts = [
        idea["hook"].strip(),
        "",
        idea["why_end"].strip(),
    ]
    if idea["twist"]:
        description_parts += ["", f"Payoff: {idea['twist'].strip()}"]
    description_parts += [
        "",
        "Educational breakdown of how this actually works -- not a how-to. "
        "If something here sounds like your situation, that's the point: "
        "knowing the mechanism is how you spot it next time.",
        "",
        "#Shorts",
    ]
    description = "\n".join(p for p in description_parts if p is not None)

    tags = [t.strip() for t in idea["category"].split("/") if t.strip()]
    tags += [t.strip() for t in idea["search_intent"].split(",") if t.strip()]

    return {
        "snippet": {
            "title": title[:100],
            "description": description[:5000],
            "tags": tags[:15],
            "categoryId": category_id,
        },
        "status": {
            "privacyStatus": None,  # set by caller
            "selfDeclaredMadeForKids": False,
        },
    }


def get_youtube_client(client_secrets, token_path):
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(client_secrets, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w") as f:
            f.write(creds.to_json())
    return build("youtube", "v3", credentials=creds)


def upload_video(youtube, video_path, body, thumbnail_path, dry_run):
    print(f"  video: {video_path}")
    print(f"  title: {body['snippet']['title']}")
    print(f"  tags:  {body['snippet']['tags']}")
    if dry_run:
        print("  [dry-run] skipping actual upload")
        return "DRY-RUN-ID"

    from googleapiclient.http import MediaFileUpload

    media = MediaFileUpload(video_path, chunksize=-1, resumable=True, mimetype="video/*")
    request = youtube.videos().insert(part="snippet,status", body=body, media_body=media)
    response = None
    while response is None:
        status, response = request.next_chunk()
        if status:
            print(f"  upload progress: {int(status.progress() * 100)}%")
    video_id = response["id"]

    if thumbnail_path:
        youtube.thumbnails().set(videoId=video_id, media_body=thumbnail_path).execute()
        print(f"  thumbnail set: {thumbnail_path}")

    return video_id


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--sheet", required=True, help="Path to the content-system .xlsx")
    p.add_argument("--video-dir", required=True, help="Folder containing rendered videos")
    p.add_argument("--client-secrets", default="client_secret.json")
    p.add_argument("--token", default="token.json")
    p.add_argument("--privacy", choices=["public", "unlisted", "private"], default="private")
    p.add_argument("--category-id", default="27", help="YouTube category ID (27 = Education)")
    p.add_argument("--dry-run", action="store_true", help="Do everything except call the YouTube API")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.sheet)
    ws = wb["Content System"] if "Content System" in wb.sheetnames else wb.active
    ensure_tracking_columns(ws)

    rows = load_rows(ws)
    candidate = None
    video_path = None
    for idea in rows:
        if idea["status"] == "Posted":
            continue
        vp = find_video_file(args.video_dir, idea["rank"], idea["title"], idea["video_file_override"])
        if vp:
            candidate = idea
            video_path = vp
            break

    if not candidate:
        print("No unposted idea has a matching video file ready in --video-dir. Nothing to post today.")
        sys.exit(1)

    print(f"Posting rank #{candidate['rank']}: {candidate['title']}")
    body = build_metadata(candidate, args.category_id)
    body["status"]["privacyStatus"] = args.privacy
    thumb = find_thumbnail(args.video_dir, candidate["rank"], candidate["title"])

    if args.dry_run:
        video_id = upload_video(None, video_path, body, thumb, dry_run=True)
    else:
        youtube = get_youtube_client(args.client_secrets, args.token)
        video_id = upload_video(youtube, video_path, body, thumb, dry_run=False)

    ws.cell(candidate["row"], STATUS_COL, "DRY-RUN" if args.dry_run else "Posted")
    ws.cell(candidate["row"], POSTED_DATE_COL, datetime.datetime.utcnow().isoformat(timespec="seconds"))
    ws.cell(candidate["row"], VIDEO_ID_COL, video_id)
    wb.save(args.sheet)

    print(f"Done. Video ID: {video_id}")


if __name__ == "__main__":
    main()
